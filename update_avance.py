"""
Actualización de avance_semanal.xlsx — Corte 10/04/2026
"""
import pandas as pd
import os
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX     = os.path.join(BASE_DIR, "data", "avance_semanal.xlsx")
XLSX_TMP = os.path.join(BASE_DIR, "data", "avance_semanal_new.xlsx")

df = pd.read_excel(XLSX)
for col in ["pct_completado", "estado", "comentario"]:
    if col not in df.columns:
        df[col] = None
    df[col] = df[col].astype(object)

# ── UPDATES ────────────────────────────────────────────────────────────────
# (planta, id_tarea, pct_completado|None, estado, comentario)
UPDATES = [
    # ── MAJES ──────────────────────────────────────────────────────────────
    ("Majes", 107, 100.0, "Completado",  "Completado ✅"),
    ("Majes", 108, 100.0, "Completado",  "Completado ✅"),
    ("Majes", 109, 100.0, "Completado",  "Completado ✅"),
    ("Majes",  78,   0.0, "En riesgo",   "Retraso +11d — plan 30/03, pendiente al 10/04"),
    ("Majes",  87,   0.0, "En riesgo",   "Retraso +11d — plan 30/03, pendiente al 10/04"),
    ("Majes", 118,   0.0, "No iniciado", "Retraso — materiales recién llegaron a obra"),
    ("Majes", 119,   0.0, "No iniciado", "Sin iniciar"),
    ("Majes", 120,   0.0, "No iniciado", "Sin iniciar"),
    ("Majes", 122,   0.0, "No iniciado", "Sin iniciar"),
    ("Majes", 131,   0.0, "En riesgo",   "🔴 BLOQUEADO — archivos DIgSILENT incorrectos de Itechene"),
    ("Majes", 133,   0.0, "En riesgo",   "🔴 BLOQUEADO — archivos DIgSILENT incorrectos de Itechene"),
    ("Majes", 134,   0.0, "No iniciado", "🔴 BLOQUEADO — depende de ID 133, sin insumos de Itechene"),
    # ── REPARTICIÓN ────────────────────────────────────────────────────────
    ("Repartición", 107, 100.0, "Completado",  "Completado 26/03 ✅"),
    ("Repartición", 108, 100.0, "Completado",  "Completado 27/03 ✅"),
    ("Repartición", 109, 100.0, "Completado",  "Completado 28/03 ✅"),
    ("Repartición", 110, 100.0, "Completado",  "Completado ✅"),
    ("Repartición",  78,   0.0, "En riesgo",   "Retraso +11d — plan 30/03, pendiente al 10/04"),
    ("Repartición",  87,   0.0, "En riesgo",   "Retraso +11d — plan 30/03, pendiente al 10/04"),
    ("Repartición", 118,   0.0, "No iniciado", "Retraso — materiales recién llegaron a obra"),
    ("Repartición", 119,   0.0, "No iniciado", "Sin iniciar"),
    ("Repartición", 120,   0.0, "No iniciado", "Sin iniciar"),
    ("Repartición", 122,   0.0, "No iniciado", "Sin iniciar"),
    ("Repartición", 131,   0.0, "En riesgo",   "🔴 BLOQUEADO — archivos DIgSILENT incorrectos de Itechene"),
    ("Repartición", 133,   0.0, "En riesgo",   "🔴 BLOQUEADO — archivos DIgSILENT incorrectos de Itechene"),
    ("Repartición", 134,   0.0, "No iniciado", "🔴 BLOQUEADO — depende de ID 133, sin insumos de Itechene"),
]

cambios = []
for (planta, id_t, pct, estado, comentario) in UPDATES:
    mask = (df["planta"] == planta) & (df["id_tarea"] == id_t)
    if mask.sum() == 0:
        print(f"  ⚠  No encontrado: {planta} ID {id_t}")
        continue
    if pct is not None:
        df.loc[mask, "pct_completado"] = pct
    df.loc[mask, "estado"]     = estado
    df.loc[mask, "comentario"] = comentario
    cambios.append(f"  ✓  {planta:12} ID {id_t:3}  pct={str(pct) if pct is not None else 'din':6}  {estado}")

print("Actualizando avance_semanal.xlsx — Corte 10/04/2026:")
for c in cambios: print(c)

# ── Guardar ────────────────────────────────────────────────────────────────
with pd.ExcelWriter(XLSX_TMP, engine="openpyxl", date_format="DD/MM/YYYY") as writer:
    df.to_excel(writer, index=False, sheet_name="Avance")
    ws = writer.sheets["Avance"]
    col_widths = {"A":14,"B":10,"C":50,"D":30,"E":18,"F":18,"G":14,"H":20,"I":18,"J":20,"K":60}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    hdr_fill = PatternFill("solid", fgColor="1A3A5C")
    hdr_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    for cell in ws[1]:
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    fill_m   = PatternFill("solid", fgColor="EBF5FB")
    fill_r   = PatternFill("solid", fgColor="EAFAF1")
    fill_upd = PatternFill("solid", fgColor="FFF3CD")
    fill_blk = PatternFill("solid", fgColor="FDECEA")  # rojo claro para BLOQUEADO
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    updated_pairs  = {(p, i) for (p, i, *_) in UPDATES}
    blocked_pairs  = {(p, i) for (p, i, _, st, _) in UPDATES if "BLOQUEADO" in _}

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        planta_val = row[0].value
        id_val     = row[1].value
        pair       = (planta_val, id_val)
        com_val    = row[10].value if len(row) > 10 else ""
        is_blocked = pair in blocked_pairs or (com_val and "BLOQUEADO" in str(com_val))
        is_updated = pair in updated_pairs

        if is_blocked:    fill = fill_blk
        elif is_updated:  fill = fill_upd
        elif planta_val == "Majes": fill = fill_m
        else:             fill = fill_r

        for cell in row:
            cell.fill   = fill; cell.border = border
            cell.font   = Font(name="Calibri", size=10, bold=is_updated)
            cell.alignment = Alignment(vertical="center")
        for ci in [1, 4, 5, 6, 7, 8]:
            row[ci].alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"

# Reemplazar original si es posible, si no dejar _new
try:
    if os.path.exists(XLSX):
        os.remove(XLSX)
    os.rename(XLSX_TMP, XLSX)
    saved = XLSX
except PermissionError:
    saved = XLSX_TMP
    print(f"  ⚠  El original sigue abierto en Excel — guardado en: avance_semanal_new.xlsx")
    print(f"     Cierra Excel y renombra manualmente, o corre de nuevo.")

print(f"\n  Guardado: {os.path.basename(saved)} ({len(df)} filas)")
print(f"  Filas actualizadas: {len(cambios)}")
