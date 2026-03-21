"""
Genera data/avance_semanal.xlsx con las 308 tareas reales
(154 Majes + 154 Repartición) directamente desde los cronogramas Excel.
"""
import pandas as pd
import os
from datetime import datetime

BASE_DIR  = os.path.dirname(os.path.abspath(__file__))
DATA_DIR  = os.path.join(BASE_DIR, "data")
FECHA_CORTE = datetime(2026, 3, 19)

MACROFASES = [
    {"id": 1, "nombre": "Ingeniería de Detalle",        "min": 3,   "max": 19},
    {"id": 2, "nombre": "Permisos HSE y Arqueológicos", "min": 20,  "max": 39},
    {"id": 3, "nombre": "Construcción Civil",           "min": 41,  "max": 71},
    {"id": 4, "nombre": "Zanjas y Cableado",            "min": 72,  "max": 95},
    {"id": 5, "nombre": "Suministro y Montaje",         "min": 96,  "max": 127},
    {"id": 6, "nombre": "Estudio COES",                 "min": 128, "max": 143},
    {"id": 7, "nombre": "Puesta en Marcha",             "min": 144, "max": 154},
]

_MESES = {
    "enero":"01","febrero":"02","marzo":"03","abril":"04",
    "mayo":"05","junio":"06","julio":"07","agosto":"08",
    "septiembre":"09","octubre":"10","noviembre":"11","diciembre":"12"
}

def parse_fecha(s):
    if pd.isna(s):
        return pd.NaT
    s = str(s).strip()
    if s.upper() in ("NOD", ""):
        return pd.NaT
    s_low = s.lower()
    for mes, num in _MESES.items():
        if mes in s_low:
            s_low = s_low.replace(mes, num)
            break
    parts = s_low.split()
    date_parts = [p for p in parts if p.isdigit() or p in _MESES.values()][:3]
    if len(date_parts) == 3:
        result = pd.to_datetime(" ".join(date_parts), dayfirst=True, errors="coerce")
        if pd.notna(result):
            return result
    return pd.to_datetime(s_low, dayfirst=True, errors="coerce")

def get_macrofase(id_val):
    for mf in MACROFASES:
        if mf["min"] <= id_val <= mf["max"]:
            return mf["nombre"]
    return "Otro"

def pct_planeado(inicio, fin, corte):
    """Porcentaje teórico completado a fecha_corte según duración planificada."""
    if pd.isna(inicio) or pd.isna(fin):
        return 0.0
    inicio = pd.Timestamp(inicio)
    fin    = pd.Timestamp(fin)
    corte  = pd.Timestamp(corte)
    if corte <= inicio:
        return 0.0
    if corte >= fin:
        return 100.0
    dur = (fin - inicio).days
    if dur == 0:
        return 100.0
    return round(100.0 * (corte - inicio).days / dur, 1)

def load_plant(filepath, planta):
    df = pd.read_excel(filepath)
    df.columns = [c.strip() for c in df.columns]

    # Normalizar nombres de columna
    col_map = {}
    for c in df.columns:
        norm = (c.replace("ó","o").replace("ú","u").replace("é","e")
                 .replace("í","i").replace("á","a").replace("ñ","n")
                 .replace(" ","_").replace("-","_"))
        col_map[c] = norm
    df = df.rename(columns=col_map)

    # Detectar columnas correctas
    nombre_col  = next((c for c in df.columns if "Nombre" in c), None)
    inicio_col  = next((c for c in df.columns if c == "Comienzo"), None)
    fin_col     = next((c for c in df.columns if c == "Fin"), None)

    rows = []
    for _, r in df.iterrows():
        id_val = r.get("Id")
        if pd.isna(id_val):
            continue
        id_val = int(id_val)

        nombre     = str(r[nombre_col]).strip() if nombre_col else ""
        inicio_raw = r[inicio_col] if inicio_col else None
        fin_raw    = r[fin_col]    if fin_col    else None

        inicio = parse_fecha(inicio_raw)
        fin    = parse_fecha(fin_raw)
        macro  = get_macrofase(id_val)
        pct_pl = pct_planeado(inicio, fin, FECHA_CORTE)

        rows.append({
            "planta":               planta,
            "id_tarea":             id_val,
            "nombre_tarea":         nombre,
            "macrofase":            macro,
            "fecha_inicio_plan":    inicio.date() if pd.notna(inicio) else None,
            "fecha_fin_plan":       fin.date()    if pd.notna(fin)    else None,
            "fecha_corte":          FECHA_CORTE.date(),
            "pct_planeado_a_hoy":   pct_pl,
            "pct_completado":       None,
            "estado":               None,
            "comentario":           None,
        })

    return pd.DataFrame(rows)

print("Leyendo Majes...")
df_m = load_plant(
    os.path.join(BASE_DIR, "Cronograma de Obra Actualizado BESS MAJES 19.03.2026.xlsx"),
    "Majes"
)
print("Leyendo Repartición...")
df_r = load_plant(
    os.path.join(BASE_DIR, "Cronograma de Obra Actualizado BESS REPARTICIÓN 19.3.26.xlsx"),
    "Repartición"
)

df_out = pd.concat([df_m, df_r], ignore_index=True)
df_out = df_out.sort_values(["planta", "id_tarea"]).reset_index(drop=True)

# ── Formatear Excel con anchos y estilos ──────────────────
out_path = os.path.join(DATA_DIR, "avance_semanal.xlsx")
with pd.ExcelWriter(out_path, engine="openpyxl", date_format="DD/MM/YYYY") as writer:
    df_out.to_excel(writer, index=False, sheet_name="Avance")

    ws = writer.sheets["Avance"]
    from openpyxl.styles import (PatternFill, Font, Alignment,
                                  Border, Side, numbers)
    from openpyxl.utils import get_column_letter

    # Anchos de columna
    col_widths = {
        "A": 14,   # planta
        "B": 10,   # id_tarea
        "C": 50,   # nombre_tarea
        "D": 30,   # macrofase
        "E": 18,   # fecha_inicio_plan
        "F": 18,   # fecha_fin_plan
        "G": 14,   # fecha_corte
        "H": 20,   # pct_planeado_a_hoy
        "I": 18,   # pct_completado
        "J": 16,   # estado
        "K": 40,   # comentario
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # Estilo encabezado
    hdr_fill = PatternFill("solid", fgColor="1A3A5C")
    hdr_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Colores por planta
    fill_m = PatternFill("solid", fgColor="EBF5FB")   # azul claro — Majes
    fill_r = PatternFill("solid", fgColor="EAFAF1")   # verde claro — Repartición
    fill_h = PatternFill("solid", fgColor="FEF9E7")   # amarillo — macrofase header

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        planta_val = row[0].value
        id_val_cell = row[1].value
        macro_val   = row[3].value

        # Detectar filas de cabecera de macrofase (id_tarea coincide con id de macrofase?
        # No — colorear solo por planta)
        fill = fill_m if planta_val == "Majes" else fill_r

        for cell in row:
            cell.fill   = fill
            cell.border = border
            cell.font   = Font(name="Calibri", size=10)
            cell.alignment = Alignment(vertical="center")

        # Centrar columnas numéricas/fecha
        for col_idx in [1, 4, 5, 6, 7, 8]:   # id, inicio, fin, corte, pct_plan, pct_real
            row[col_idx].alignment = Alignment(horizontal="center", vertical="center")

    # Congelar primera fila
    ws.freeze_panes = "A2"

    # Formato porcentaje para columnas H e I
    pct_fmt = '0.0"%"'
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=8, max_col=9):
        for cell in row:
            if cell.value is not None:
                cell.number_format = pct_fmt

print(f"\nGenerado: data/avance_semanal.xlsx")
print(f"  Filas: {len(df_out)}  ({len(df_m)} Majes + {len(df_r)} Repartición)")
print(f"  Columnas: {list(df_out.columns)}")

# Validación rápida
print(f"\nMuestra pct_planeado_a_hoy:")
for mf in MACROFASES:
    sub = df_out[df_out["macrofase"] == mf["nombre"]]
    pct_m = sub[sub["planta"]=="Majes"]["pct_planeado_a_hoy"].mean()
    pct_r = sub[sub["planta"]=="Repartición"]["pct_planeado_a_hoy"].mean()
    print(f"  {mf['nombre']:<32} Majes={pct_m:5.1f}%  Rep={pct_r:5.1f}%")

print(f"\nFechas parseadas (sample Majes):")
sample = df_out[df_out["planta"]=="Majes"].head(5)[
    ["id_tarea","nombre_tarea","fecha_inicio_plan","fecha_fin_plan","pct_planeado_a_hoy"]]
print(sample.to_string(index=False))
